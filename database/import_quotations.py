#!/usr/bin/env python3
"""
PEGASUS ERP - Quotation Import Script
1. Delete all existing quotation data
2. Read Excel quotation_list.xlsx, group by AA column
3. Match PDF files to extract line items
4. Insert into quotation_headers + quotation_lines
"""

import os
import sys
import re
import json
import psycopg2
import openpyxl
from datetime import datetime, date

# ── Config ──
EXCEL_PATH = r"C:\Users\R.Nozaki\Downloads\quotation_list.xlsx"
QT_DATA_DIR = r"C:\Users\R.Nozaki\Downloads\Pegasus_ERP_R1\QT data"
DB_CONFIG = {
    'host': 'localhost',
    'port': '5432',
    'dbname': 'pegasus_erp',
    'user': 'postgres',
    'password': 'postgres'
}

try:
    import fitz  # pymupdf
    HAS_FITZ = True
except ImportError:
    HAS_FITZ = False
    print("WARNING: pymupdf not installed, skipping PDF parsing")


def parse_date(val):
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.strftime('%Y-%m-%d') if val.year > 1900 else None
    s = str(val).strip()
    if not s or s == 'None':
        return None
    try:
        return datetime.strptime(s[:10], '%Y-%m-%d').strftime('%Y-%m-%d')
    except:
        return None


def parse_float(val):
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(',', '')
    try:
        return float(s)
    except:
        return 0.0


def clean_str(val):
    if val is None:
        return None
    s = str(val).strip()
    return s if s and s != 'None' else None


def parse_possibility(val):
    if not val:
        return 'DRAFT', None
    s = str(val).strip()
    if 'Deposited' in s or 'Invoice' in s:
        return 'WON', s
    if 'PO' in s and 'PO Next' not in s:
        return 'WON', s
    if 'Lost' in s:
        return 'LOST', s
    if 'Revised' in s:
        return 'CANCELLED', s
    if 'Fix price' in s or 'A:' in s:
        return 'SUBMITTED', s
    if 'C:' in s or 'D:' in s:
        return 'SUBMITTED', s
    if 'E:' in s:
        return 'DRAFT', s
    if 'Pending' in s:
        return 'DRAFT', s
    return 'DRAFT', s


def build_pdf_index(qt_data_dir):
    """Build index: filename -> full_path for all PDFs"""
    pdf_index = {}
    for root, dirs, files in os.walk(qt_data_dir):
        for f in files:
            if f.lower().endswith('.pdf'):
                pdf_index[f] = os.path.join(root, f)
    print(f"  Found {len(pdf_index)} PDF files")
    return pdf_index


def find_matching_pdf(quotation_no, customer_name, pdf_index):
    """Find PDF matching quotation_no pattern in filename"""
    if not quotation_no:
        return None

    qt_norm = quotation_no.strip()
    # QT number in filename uses _ not -, e.g. QT260403_01 for QT260403-01
    qt_file_prefix = qt_norm.replace('-', '_')

    for pdf_name, pdf_path in pdf_index.items():
        if pdf_name.startswith(qt_file_prefix) or pdf_name.startswith(qt_norm):
            return pdf_path

    return None


def extract_pdf_lines(pdf_path):
    """Extract line items from a TOMAS TECH quotation PDF"""
    if not HAS_FITZ:
        return [], {}

    try:
        doc = fitz.open(pdf_path)
    except Exception as e:
        return [], {}

    all_lines = []
    header_info = {}

    for page_num in range(len(doc)):
        page = doc[page_num]

        # Extract header info from text blocks (first page only)
        if page_num == 0:
            blocks = page.get_text('dict')['blocks']
            for b in blocks:
                if 'lines' not in b:
                    continue
                for line in b['lines']:
                    text = ''.join([span['text'] for span in line['spans']])
                    bbox = line['bbox']
                    if text.strip() and bbox[1] < 330:
                        x, y = bbox[0], bbox[1]
                        if 140 < x < 500 and 205 < y < 220:
                            header_info['customer'] = text.strip()
                        if 140 < x < 500 and 220 < y < 240:
                            header_info['attention'] = text.strip()
                        if 140 < x < 600 and 268 < y < 282:
                            header_info['project'] = text.strip()
                        if 680 < x < 800 and 268 < y < 282:
                            header_info['quoted_by'] = text.strip()

        # Extract table
        tables = page.find_tables()
        for tab in tables.tables:
            data = tab.extract()
            if len(data) < 2:
                continue

            header_row = data[0]
            if not header_row or not any('Item' in str(c) or 'Description' in str(c) for c in header_row if c):
                continue

            # Items packed in row 1
            if len(data) > 1 and data[1]:
                items_row = data[1]
                nos_str = items_row[0] or ''
                desc_str = items_row[1] or ''
                qty_str = items_row[2] or ''
                unit_str = items_row[3] or ''
                price_str = items_row[4] or ''
                ext_str = items_row[5] or ''

                nos = [x.strip() for x in nos_str.split('\n') if x.strip()]
                descs = [x.strip() for x in desc_str.split('\n') if x.strip()]
                qtys = [x.strip() for x in qty_str.split('\n') if x.strip()]
                units = [x.strip() for x in unit_str.split('\n') if x.strip()]
                prices = [x.strip() for x in price_str.split('\n') if x.strip()]
                exts = [x.strip() for x in ext_str.split('\n') if x.strip()]

                # Map descriptions to line numbers
                # Strategy: walk through nos, for each no consume descriptions
                desc_idx = 0
                qty_idx = 0
                unit_idx = 0
                price_idx = 0
                ext_idx = 0

                for no_idx, no in enumerate(nos):
                    is_cat = bool(re.match(r'^\d+$', no))
                    is_item = bool(re.match(r'^\d+-\d+$', no))

                    if is_cat:
                        cat_desc = descs[desc_idx] if desc_idx < len(descs) else ''
                        desc_idx += 1
                        all_lines.append({
                            'line_no': no,
                            'is_category': True,
                            'description': cat_desc,
                            'quantity': None,
                            'unit': None,
                            'unit_price': 0,
                            'ext_price': 0,
                        })
                    elif is_item:
                        # Gather description lines until next no entry
                        item_desc_parts = []
                        if desc_idx < len(descs):
                            item_desc_parts.append(descs[desc_idx])
                            desc_idx += 1

                        # Check if there are extra desc lines before next numbered line
                        next_no_idx_pos = no_idx + 1
                        if next_no_idx_pos < len(nos):
                            # Peek ahead to see how many descs before next no
                            while desc_idx < len(descs):
                                # If next desc looks like it belongs to next no, stop
                                # Heuristic: check if next no's category desc matches
                                next_no = nos[next_no_idx_pos] if next_no_idx_pos < len(nos) else None
                                if next_no and re.match(r'^\d+$', next_no):
                                    # Next is category - check if desc matches category pattern
                                    break
                                elif next_no and re.match(r'^\d+-\d+$', next_no):
                                    break
                                # Extra description line for current item
                                item_desc_parts.append(descs[desc_idx])
                                desc_idx += 1

                        qty_val = None
                        unit_val = None
                        price_val = 0
                        ext_val = 0

                        if qty_idx < len(qtys):
                            try:
                                qty_val = float(qtys[qty_idx].replace(',', ''))
                            except:
                                qty_val = 1
                            qty_idx += 1

                        if unit_idx < len(units):
                            unit_val = units[unit_idx]
                            unit_idx += 1

                        if price_idx < len(prices):
                            try:
                                price_val = float(prices[price_idx].replace(',', ''))
                            except:
                                price_val = 0
                            price_idx += 1

                        if ext_idx < len(exts):
                            try:
                                ext_val = float(exts[ext_idx].replace(',', ''))
                            except:
                                ext_val = 0
                            ext_idx += 1

                        all_lines.append({
                            'line_no': no,
                            'is_category': False,
                            'description': ' / '.join(item_desc_parts) if item_desc_parts else '',
                            'quantity': qty_val,
                            'unit': unit_val,
                            'unit_price': price_val,
                            'ext_price': ext_val,
                        })

            # Extract totals
            for row in data[2:]:
                if row and row[3]:
                    label = str(row[3]).strip()
                    val_str = str(row[5] or '').strip().replace(',', '')
                    try:
                        val = float(val_str)
                    except:
                        val = 0
                    if 'Sub Total' in label:
                        header_info['subtotal'] = val
                    elif 'Discount' in label:
                        header_info['discount'] = val
                    elif 'Sales Amount' in label:
                        header_info['sales_amount'] = val
                    elif 'Value Added Tax' in label:
                        header_info['vat'] = val
                    elif 'Grand Total' in label:
                        header_info['grand_total'] = val

        # Remark, lead time
        text = page.get_text()
        remark_match = re.search(r'Remark\s*:\s*\n(.+?)(?=Currency|$)', text, re.DOTALL)
        if remark_match:
            remark_lines = [l.strip() for l in remark_match.group(1).strip().split('\n') if l.strip()]
            header_info['remark'] = '\n'.join(remark_lines[:5])

        lt_match = re.search(r'Lead\s*time\s*:\s*(.+)', text)
        if lt_match:
            header_info['lead_time'] = lt_match.group(1).strip()

    doc.close()
    return all_lines, header_info


def main():
    print("=" * 60)
    print("PEGASUS ERP - Quotation Import")
    print("=" * 60)

    conn = psycopg2.connect(**DB_CONFIG)
    conn.autocommit = False
    cur = conn.cursor()

    # ── Step 1: Delete existing quotation data ──
    print("\n[1] Deleting existing quotation data...")
    # Clear FK references first
    cur.execute("UPDATE sales_order_headers SET quotation_id = NULL WHERE quotation_id IS NOT NULL")
    cur.execute("UPDATE purchase_order_headers SET our_quotation_id = NULL WHERE our_quotation_id IS NOT NULL")
    cur.execute("UPDATE ar_invoices SET our_quotation_id = NULL WHERE our_quotation_id IS NOT NULL")
    cur.execute("UPDATE cost_sheets SET quotation_id = NULL WHERE quotation_id IS NOT NULL")
    cur.execute("UPDATE project_cost_items SET quotation_id = NULL WHERE quotation_id IS NOT NULL")
    cur.execute("DELETE FROM quotation_lines")
    cur.execute("DELETE FROM quotation_headers")
    conn.commit()
    print("  Done - all quotation data deleted")

    # ── Step 2: Read Excel ──
    print("\n[2] Reading Excel file...")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb.active

    # Build customer lookup
    cur.execute("SELECT customer_id, customer_name FROM customers WHERE is_deleted = FALSE")
    customer_map = {}
    for row in cur.fetchall():
        customer_map[row[1].strip()] = row[0]
    print(f"  Loaded {len(customer_map)} customers from DB")

    # Build solution category lookup
    cur.execute("SELECT category_id, category_name FROM solution_categories WHERE is_deleted = FALSE")
    solution_cat_map = {}
    for row in cur.fetchall():
        solution_cat_map[row[1].strip()] = row[0]
    print(f"  Loaded {len(solution_cat_map)} solution categories from DB")

    # Group rows by AA column
    groups = {}
    skipped = 0
    for r in range(3, ws.max_row + 1):
        aa = ws.cell(r, 27).value
        if aa is None:
            continue
        aa = str(aa).strip()
        if not aa or aa.startswith('-'):
            skipped += 1
            continue

        row_data = {
            'date': ws.cell(r, 2).value,
            'quotation_no': clean_str(ws.cell(r, 3).value),
            'pj': clean_str(ws.cell(r, 4).value),
            'related_projects': clean_str(ws.cell(r, 5).value),
            'po_date': ws.cell(r, 6).value,
            'customer': clean_str(ws.cell(r, 7).value),
            'customer_staff': clean_str(ws.cell(r, 8).value),
            'touch_point': clean_str(ws.cell(r, 9).value),
            'solution': clean_str(ws.cell(r, 10).value),
            'solution_category': clean_str(ws.cell(r, 11).value),
            'title': clean_str(ws.cell(r, 12).value),
            'sales_amount': parse_float(ws.cell(r, 13).value),
            'service_cost': parse_float(ws.cell(r, 14).value),
            'engineer_cost': parse_float(ws.cell(r, 15).value),
            'commission': parse_float(ws.cell(r, 16).value),
            'service_profit': parse_float(ws.cell(r, 17).value),
            'service_profit_pct': parse_float(ws.cell(r, 18).value),
            'gross_profit': parse_float(ws.cell(r, 19).value),
            'gross_profit_pct': parse_float(ws.cell(r, 20).value),
            'possibility': clean_str(ws.cell(r, 21).value),
            'invoice_schedule': clean_str(ws.cell(r, 22).value),
            'income_schedule': clean_str(ws.cell(r, 23).value),
            'budget': clean_str(ws.cell(r, 24).value),
            'sales_name': clean_str(ws.cell(r, 25).value),
            'note': clean_str(ws.cell(r, 26).value),
        }

        if aa not in groups:
            groups[aa] = []
        groups[aa].append(row_data)

    print(f"  Total data rows: {sum(len(v) for v in groups.values())}")
    print(f"  Skipped rows: {skipped}")
    print(f"  Unique quotations (AA groups): {len(groups)}")

    # ── Step 3: Build PDF index ──
    print("\n[3] Building PDF index...")
    pdf_index = build_pdf_index(QT_DATA_DIR)

    # ── Step 4: Insert quotations ──
    print("\n[4] Inserting quotations...")
    inserted = 0
    pdf_matched = 0
    lines_inserted = 0
    errors = []

    for aa_key, rows in groups.items():
        try:
            cur.execute("SAVEPOINT sp_qt")
            first = rows[0]

            # Sum amounts across all rows in group
            total_sales = sum(r['sales_amount'] for r in rows)
            total_service_cost = sum(r['service_cost'] for r in rows)
            total_engineer_cost = sum(r['engineer_cost'] for r in rows)
            total_commission = sum(r['commission'] for r in rows)
            total_service_profit = sum(r['service_profit'] for r in rows)
            total_gross_profit = sum(r['gross_profit'] for r in rows)
            avg_service_profit_pct = (total_service_profit / total_sales * 100) if total_sales > 0 else 0
            avg_gross_profit_pct = (total_gross_profit / total_sales * 100) if total_sales > 0 else 0

            qt_no_base = first['quotation_no'] or f"IMPORT-{inserted+1:04d}"
            # Ensure unique quotation_no - check for duplicates
            qt_no = qt_no_base
            cur.execute("SELECT 1 FROM quotation_headers WHERE quotation_no = %s", [qt_no])
            if cur.fetchone():
                suffix = 2
                while True:
                    qt_no = f"{qt_no_base}_{suffix}"
                    cur.execute("SELECT 1 FROM quotation_headers WHERE quotation_no = %s", [qt_no])
                    if not cur.fetchone():
                        break
                    suffix += 1

            # Customer lookup / create
            customer_name = first['customer']
            customer_id = None
            if customer_name:
                cn = customer_name.strip()
                customer_id = customer_map.get(cn)
                if not customer_id:
                    for cname, cid in customer_map.items():
                        if cn.lower() == cname.lower():
                            customer_id = cid
                            break
                if not customer_id:
                    cur.execute(
                        "INSERT INTO customers (customer_name, customer_code, is_current, is_deleted) "
                        "VALUES (%s, %s, TRUE, FALSE) RETURNING customer_id",
                        [cn, qt_no[:3].upper()]
                    )
                    customer_id = cur.fetchone()[0]
                    customer_map[cn] = customer_id

            # Solution category lookup / create
            solution_cat_id = None
            sol_cat_name = first['solution_category']
            if sol_cat_name:
                sc = sol_cat_name.strip()
                solution_cat_id = solution_cat_map.get(sc)
                if not solution_cat_id:
                    for sc_name, sc_id in solution_cat_map.items():
                        if sc in sc_name or sc_name in sc:
                            solution_cat_id = sc_id
                            break
                    if not solution_cat_id:
                        cur.execute(
                            "INSERT INTO solution_categories (category_name, sort_order, is_deleted) "
                            "VALUES (%s, 99, FALSE) RETURNING category_id",
                            [sc]
                        )
                        solution_cat_id = cur.fetchone()[0]
                        solution_cat_map[sc] = solution_cat_id

            issue_date = parse_date(first['date']) or '2024-01-01'
            po_date = parse_date(first['po_date'])
            status, possibility_text = parse_possibility(first['possibility'])

            titles = [r['title'] for r in rows if r['title']]
            combined_title = titles[0] if titles else qt_no
            notes = [r['note'] for r in rows if r['note']]
            combined_notes = '; '.join(set(notes)) if notes else None

            vat_rate = 7.0
            subtotal = total_sales
            vat_amount = round(subtotal * vat_rate / 100, 2)
            grand_total = round(subtotal + vat_amount, 2)

            invoice_schedule = first.get('invoice_schedule')
            income_schedule = first.get('income_schedule')

            cur.execute("""
                INSERT INTO quotation_headers (
                    quotation_no, issue_date, customer_id, division_id,
                    currency_code, exchange_rate,
                    subtotal_thb, discount_amount, vat_rate, vat_amount, grand_total_thb,
                    project_name, project_code, attention_name,
                    remark_text, note_text, status, possibility,
                    solution_category_id, solution_name,
                    pj_no, customer_staff, touch_point,
                    service_cost, engineer_cost, commission,
                    service_profit, service_profit_pct,
                    gross_profit, gross_profit_pct,
                    invoice_schedule, income_schedule, budget,
                    sales_name, unique_key, po_date,
                    created_by, created_at
                ) VALUES (
                    %s, %s, %s, 1,
                    'THB', 1,
                    %s, 0, %s, %s, %s,
                    %s, %s, %s,
                    %s, %s, %s, %s,
                    %s, %s,
                    %s, %s, %s,
                    %s, %s, %s,
                    %s, %s,
                    %s, %s,
                    %s, %s, %s,
                    %s, %s, %s,
                    1, NOW()
                ) RETURNING quotation_id
            """, [
                qt_no, issue_date, customer_id,
                subtotal, vat_rate, vat_amount, grand_total,
                combined_title, first['pj'],
                first['customer_staff'],
                None, combined_notes, status, possibility_text,
                solution_cat_id, first['solution'],
                first['pj'], first['customer_staff'], first['touch_point'],
                total_service_cost, total_engineer_cost, total_commission,
                total_service_profit, avg_service_profit_pct,
                total_gross_profit, avg_gross_profit_pct,
                invoice_schedule, income_schedule, first['budget'],
                first['sales_name'], aa_key, po_date,
            ])
            quotation_id = cur.fetchone()[0]
            inserted += 1

            # ── Try PDF matching ──
            pdf_path = find_matching_pdf(qt_no, customer_name, pdf_index)
            if pdf_path:
                pdf_matched += 1
                pdf_lines, pdf_header = extract_pdf_lines(pdf_path)

                if pdf_lines:
                    # Update header with PDF info
                    updates = []
                    params = []
                    if pdf_header.get('remark'):
                        updates.append("remark_text = %s")
                        params.append(pdf_header['remark'])
                    if pdf_header.get('lead_time'):
                        updates.append("lead_time_text = %s")
                        params.append(pdf_header['lead_time'])
                    if pdf_header.get('attention'):
                        updates.append("attention_name = %s")
                        params.append(pdf_header['attention'])
                    if updates:
                        params.append(quotation_id)
                        cur.execute(
                            f"UPDATE quotation_headers SET {', '.join(updates)} WHERE quotation_id = %s",
                            params
                        )

                    # Insert PDF line items
                    sort_order = 0
                    used_line_nos = set()
                    for line in pdf_lines:
                        sort_order += 1
                        is_cat = line['is_category']
                        parent_ln = None
                        ln = str(line['line_no'])
                        if not is_cat and '-' in ln:
                            parent_ln = ln.split('-')[0]

                        # Ensure unique line_no per quotation
                        orig_ln = ln
                        suffix = 1
                        while ln in used_line_nos:
                            suffix += 1
                            if '-' in orig_ln:
                                parts = orig_ln.split('-')
                                ln = f"{parts[0]}-{int(parts[1])+suffix-1}"
                            else:
                                ln = f"{orig_ln}_{suffix}"
                        used_line_nos.add(ln)

                        cur.execute("""
                            INSERT INTO quotation_lines (
                                quotation_id, line_no, parent_line_no, is_category_row,
                                item_description, quantity, unit,
                                unit_price, discount_rate, ext_price, cost_total,
                                sort_order
                            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, 0, %s, 0, %s)
                        """, [
                            quotation_id, ln, parent_ln, is_cat,
                            line['description'] or '',
                            line['quantity'], line['unit'],
                            line['unit_price'], line['ext_price'],
                            sort_order,
                        ])
                        lines_inserted += 1
                else:
                    # PDF found but no items - use Excel rows
                    sort_order = 0
                    for r_data in rows:
                        sort_order += 1
                        cur.execute("""
                            INSERT INTO quotation_lines (
                                quotation_id, line_no, parent_line_no, is_category_row,
                                item_description, quantity, unit,
                                unit_price, discount_rate, ext_price, cost_total,
                                sort_order
                            ) VALUES (%s, %s, NULL, FALSE, %s, 1, 'LOT', %s, 0, %s, %s, %s)
                        """, [
                            quotation_id, str(sort_order),
                            r_data['title'] or qt_no,
                            r_data['sales_amount'], r_data['sales_amount'],
                            r_data['service_cost'], sort_order,
                        ])
                        lines_inserted += 1
            else:
                # No PDF - use Excel rows as lines
                sort_order = 0
                for r_data in rows:
                    sort_order += 1
                    cur.execute("""
                        INSERT INTO quotation_lines (
                            quotation_id, line_no, parent_line_no, is_category_row,
                            item_description, quantity, unit,
                            unit_price, discount_rate, ext_price, cost_total,
                            sort_order
                        ) VALUES (%s, %s, NULL, FALSE, %s, 1, 'LOT', %s, 0, %s, %s, %s)
                    """, [
                        quotation_id, str(sort_order),
                        r_data['title'] or qt_no,
                        r_data['sales_amount'], r_data['sales_amount'],
                        r_data['service_cost'], sort_order,
                    ])
                    lines_inserted += 1

            if inserted % 200 == 0:
                conn.commit()
                print(f"  ... {inserted} quotations inserted")

        except Exception as e:
            cur.execute("ROLLBACK TO SAVEPOINT sp_qt")
            errors.append((aa_key, str(e)))
            if len(errors) <= 5:
                print(f"  ERROR [{aa_key[:40]}]: {e}")

    conn.commit()

    print("\n" + "=" * 60)
    print("Import Complete!")
    print(f"  Quotations inserted: {inserted}")
    print(f"  PDF files matched:   {pdf_matched}")
    print(f"  Line items inserted: {lines_inserted}")
    print(f"  Errors:              {len(errors)}")
    if errors:
        for k, e in errors[:10]:
            print(f"    {k[:50]}: {e}")
    print("=" * 60)

    cur.close()
    conn.close()


if __name__ == '__main__':
    main()

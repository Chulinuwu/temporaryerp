import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\R.Nozaki\Downloads\E-mail list.xlsx', data_only=True)
ws = wb['BUSINESS CARD']

admin_codes = {'EMP001', 'EMP057'}
manager_codes = {'EMP003', 'EMP013', 'EMP020', 'EMP025', 'EMP056'}
accounting_codes = {'EMP015'}

def clean(s, default=''):
    if s is None or str(s).strip() in ('-', ''):
        return default
    return str(s).replace('\u200b', '').replace('\u3000', ' ').replace("'", "''").strip()

def sq(s):
    return "'" + s + "'" if s else "NULL"

employees = []
for row_idx in range(7, ws.max_row + 1):
    no = ws.cell(row=row_idx, column=1).value
    if no is None or not isinstance(no, (int, float)):
        continue
    no = int(no)
    type_val = ws.cell(row=row_idx, column=23).value
    if type_val != 0:
        continue

    name_en = clean(ws.cell(row=row_idx, column=3).value)
    name_jp = clean(ws.cell(row=row_idx, column=2).value)
    name_th = clean(ws.cell(row=row_idx, column=4).value)
    if not name_en and not name_th:
        continue

    dept = clean(ws.cell(row=row_idx, column=5).value)
    job = clean(ws.cell(row=row_idx, column=6).value)
    phone = clean(ws.cell(row=row_idx, column=7).value)
    email = clean(ws.cell(row=row_idx, column=8).value)
    nickname = clean(ws.cell(row=row_idx, column=9).value)
    hire_date = ws.cell(row=row_idx, column=14).value
    salary = ws.cell(row=row_idx, column=29).value

    emp_code = f'EMP{no:03d}'
    hire_str = hire_date.strftime('%Y-%m-%d') if hire_date and hasattr(hire_date, 'strftime') else '2024-01-01'
    sal = int(salary) if salary and isinstance(salary, (int, float)) else 0

    if emp_code in admin_codes:
        role = 'ADMIN'
    elif emp_code in manager_codes:
        role = 'SALES_MANAGER'
    elif emp_code in accounting_codes:
        role = 'ACCOUNTING'
    else:
        role = 'STAFF'

    nat = 'TH'
    if emp_code in ('EMP001', 'EMP108'):
        nat = 'JP'
    elif any(x in name_en for x in ('Nguyen', 'Pham', 'Phan', 'Doan')):
        nat = 'VN'
    elif name_jp:
        for c in name_jp:
            if ord(c) >= 0x3000:
                nat = 'JP'
                break

    username = email.split('@')[0] if email else emp_code.lower()

    employees.append({
        'emp_code': emp_code, 'name_en': name_en, 'name_jp': name_jp,
        'name_th': name_th, 'dept': dept, 'job': job, 'phone': phone,
        'email': email, 'nickname': nickname, 'hire_date': hire_str,
        'salary': sal, 'nat': nat, 'role': role, 'username': username,
    })

# Generate SQL
lines = []
lines.append("-- PEGASUS ERP - Employee & User Seed Data")
lines.append("-- Generated from E-mail list.xlsx | 2026-04-14")
lines.append("-- Roles: ADMIN(2), SALES_MANAGER(5), ACCOUNTING(1), STAFF(rest)")
lines.append("")
lines.append("-- Update role constraint")
lines.append("ALTER TABLE users DROP CONSTRAINT IF EXISTS users_role_check;")
lines.append("ALTER TABLE users ADD CONSTRAINT users_role_check")
lines.append("    CHECK (role IN ('ADMIN','SALES_MANAGER','ACCOUNTING','STAFF'));")
lines.append("")

pw_hash = "$2y$10$92IXUNpkjO0rOQ5byMi.Ye4oKoEa3Ro9llC/.og/at2.uheWG/igi"

for e in employees:
    lines.append(f"-- {e['emp_code']}: {e['name_en']} [{e['role']}]")

    if e['emp_code'] == 'EMP001':
        lines.append(f"UPDATE employees SET base_salary = {e['salary']}.00, position_title = '{e['job']}', phone = {sq(e['phone'])}, email = {sq(e['email'])}, role = '{e['role']}' WHERE emp_code = 'EMP001';")
        lines.append(f"UPDATE users SET role = '{e['role']}' WHERE username = 'admin';")
        lines.append("")
        continue

    nick_sql = sq(e['nickname'])
    email_sql = sq(e['email'])
    phone_sql = sq(e['phone'])

    lines.append(f"INSERT INTO employees (emp_code, division_id, full_name, full_name_jp, full_name_th, nickname, nationality, hire_date, employment_type, position_title, salary_type, base_salary, salary_currency, email, phone, annual_leave_days, sick_leave_days, leave_balance_annual, leave_balance_sick, role)")
    lines.append(f"VALUES ('{e['emp_code']}', 1, '{e['name_en']}', '{e['name_jp']}', '{e['name_th']}', {nick_sql}, '{e['nat']}', '{e['hire_date']}', 'FULL_TIME', '{e['job']}', 'MONTHLY', {e['salary']}.00, 'THB', {email_sql}, {phone_sql}, 6, 30, 6, 30, '{e['role']}')")
    lines.append(f"ON CONFLICT (emp_code) DO UPDATE SET full_name=EXCLUDED.full_name, full_name_jp=EXCLUDED.full_name_jp, full_name_th=EXCLUDED.full_name_th, position_title=EXCLUDED.position_title, base_salary=EXCLUDED.base_salary, email=EXCLUDED.email, phone=EXCLUDED.phone, role=EXCLUDED.role;")
    lines.append("")

lines.append("")
lines.append("-- USER ACCOUNTS (default password: password123)")
lines.append("")

for e in employees:
    if e['emp_code'] == 'EMP001':
        continue
    email_sql = sq(e['email'])
    lines.append(f"INSERT INTO users (username, password_hash, email, role, employee_id, is_active)")
    lines.append(f"SELECT '{e['username']}', '{pw_hash}', {email_sql}, '{e['role']}', e.employee_id, TRUE")
    lines.append(f"FROM employees e WHERE e.emp_code = '{e['emp_code']}'")
    lines.append(f"ON CONFLICT (username) DO UPDATE SET role=EXCLUDED.role, employee_id=EXCLUDED.employee_id;")
    lines.append("")

with open(r'C:\Users\R.Nozaki\Downloads\Pegasus_ERP_R1\database\seed_employees.sql', 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines))

print(f"Generated SQL for {len(employees)} employees")
print(f"  ADMIN: {sum(1 for e in employees if e['role']=='ADMIN')}")
print(f"  SALES_MANAGER: {sum(1 for e in employees if e['role']=='SALES_MANAGER')}")
print(f"  ACCOUNTING: {sum(1 for e in employees if e['role']=='ACCOUNTING')}")
print(f"  STAFF: {sum(1 for e in employees if e['role']=='STAFF')}")

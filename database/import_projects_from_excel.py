"""
Phase 1: Import projects from the aggregated Excel file.

Reads:  QT data/Quotation_List_Aggregated_by_AA.xlsx (sheet "list (集約)")
Writes: projects table

Rules:
  - Column U (Possibility) determines whether to register:
        Deposited :100%               → status COMPLETED
        Invoice   :100%               → status ACTIVE (running, billed)
        PO        : 95%               → status ACTIVE (PO issued)
        A:Fix price & waiting PO :90% → status ACTIVE (waiting PO)
        everything else               → skip
  - Project number = column D (PJ).  If empty, auto-number PJ-AUTO-<seq>.
  - Customer name = column G.  If customer not in master, register a new one (auto-code).
  - Project name = column L (Title/Description) fallback to "PJ-XXX (auto)".
"""
from __future__ import annotations
import sys, os, re, datetime
sys.stdout.reconfigure(encoding='utf-8')

import openpyxl, psycopg
from psycopg.rows import dict_row

XLSX = r"C:\Users\R.Nozaki\Downloads\Pegasus_ERP_R1\QT data\Quotation_List_Aggregated_by_AA.xlsx"

POSSIBILITY_TO_STATUS = {
    "Deposited":  "COMPLETED",
    "Invoice":    "ACTIVE",
    "PO":         "ACTIVE",
    "A:Fix price & waiting PO": "ACTIVE",
}

def normalize_possibility(s: str) -> str | None:
    if not s: return None
    head = re.split(r"\s*:", s.strip(), maxsplit=1)[0].strip()
    return head if head in POSSIBILITY_TO_STATUS else None

def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]

    conn = psycopg.connect(
        host="localhost", port=5432, dbname="pegasus_erp",
        user="postgres", password=os.environ.get("PG_PASSWORD","postgres"),
        autocommit=False, row_factory=dict_row,
    )
    cur = conn.cursor()

    # Existing customers by normalised name
    def norm_name(n: str) -> str:
        return re.sub(r"[\s,\.\-_]+", "", (n or "")).upper()

    cust_lookup: dict[str, int] = {}
    for r in cur.execute("SELECT customer_id, customer_name FROM customers WHERE is_deleted=false AND is_current=true"):
        cust_lookup[norm_name(r["customer_name"])] = r["customer_id"]

    # Existing projects by pj_no
    pj_lookup: dict[str, int] = {
        r["pj_no"]: r["project_id"]
        for r in cur.execute("SELECT project_id, pj_no FROM projects WHERE status != 'CANCELLED'")
    }

    div_id = cur.execute("SELECT division_id FROM divisions WHERE is_deleted=false ORDER BY division_id LIMIT 1").fetchone()["division_id"]
    admin_uid = cur.execute("SELECT user_id FROM users WHERE role='ADMIN' AND is_active=true ORDER BY user_id LIMIT 1").fetchone()["user_id"]

    stats = {"considered":0, "skipped":0, "registered":0, "existing":0, "cust_created":0}
    auto_seq = 1

    print("Scanning Excel rows...")
    for i, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        if not row or len(row) < 22: continue
        date_val, qt_no, pj_no, _rel, _po_date, customer, _staff, _tp, _sol, _scat, title, *_rest = row[1:]
        # col U (Possibility) is index 20
        possibility = row[20] if len(row) > 20 else None
        status_norm = normalize_possibility(str(possibility) if possibility else "")
        if status_norm is None:
            stats["skipped"] += 1
            continue
        stats["considered"] += 1

        pj_status = POSSIBILITY_TO_STATUS[status_norm]
        pj_no_str = str(pj_no).strip() if pj_no else ""
        if not pj_no_str:
            pj_no_str = f"PJ-AUTO-{auto_seq:04d}"
            auto_seq += 1

        if pj_no_str in pj_lookup:
            stats["existing"] += 1
            continue

        # Resolve customer
        cust_name = (customer or "").strip() if customer else ""
        cust_id = None
        if cust_name:
            cust_id = cust_lookup.get(norm_name(cust_name))
            if not cust_id:
                # Create minimal customer (APPROVED so it's immediately usable)
                code = f"CUS-IMP-{stats['cust_created']+1:04d}"
                new_cust = cur.execute(
                    """
                    INSERT INTO customers
                      (customer_code, customer_name, division_id, country,
                       currency_code, payment_terms, approval_status,
                       created_by, submitted_by, submitted_at,
                       manager_approved_by, manager_approved_at,
                       ceo_approved_by,     ceo_approved_at)
                    VALUES (%s, %s, %s, 'TH', 'THB', 30, 'APPROVED',
                            %s, %s, NOW(), %s, NOW(), %s, NOW())
                    RETURNING customer_id
                    """,
                    (code, cust_name[:200], div_id, admin_uid, admin_uid, admin_uid, admin_uid),
                ).fetchone()
                cust_id = new_cust["customer_id"]
                cust_lookup[norm_name(cust_name)] = cust_id
                stats["cust_created"] += 1

        # Pick project_date
        pdate = None
        if isinstance(date_val, datetime.datetime):
            pdate = date_val.date()
        elif isinstance(date_val, datetime.date):
            pdate = date_val
        else:
            pdate = datetime.date.today()

        pname = (str(title).strip()[:480]) if title else f"{pj_no_str} (imported)"

        cur.execute(
            """
            INSERT INTO projects
              (pj_no, pj_name, customer_id, status,
               start_work_date, created_by, created_at, updated_at)
            VALUES (%s, %s, %s, %s, %s, %s, NOW(), NOW())
            RETURNING project_id
            """,
            (pj_no_str[:20], pname, cust_id, pj_status, pdate, admin_uid),
        )
        pj_lookup[pj_no_str] = cur.fetchone()["project_id"]
        stats["registered"] += 1

        if stats["registered"] % 50 == 0:
            print(f"  ...{stats['registered']} projects so far")

    conn.commit()
    print("\nDone.")
    print(f"  Rows considered  (Deposited/Invoice/PO/Fix-waiting-PO): {stats['considered']}")
    print(f"  Already existed:                                        {stats['existing']}")
    print(f"  Projects registered:                                    {stats['registered']}")
    print(f"  Customers newly created:                                {stats['cust_created']}")
    print(f"  Rows skipped (other Possibility):                       {stats['skipped']}")
    conn.close()

if __name__ == "__main__":
    main()

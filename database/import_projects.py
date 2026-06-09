#!/usr/bin/env python3
"""
PEGASUS ERP - Import projects + invoices + purchases from Excel.

Sources:
 1. quotation_list.xlsx / Base   -> projects master (one row per PJ No)
 2. quotation_list.xlsx / list   -> ensure any quotation with PJ No has a project row
 3. PJ_list_20251017_R0-1.xlsx   -> per-sheet: invoices (left) + purchases (right)
"""
import sys, io, os, re
from datetime import datetime, date
import openpyxl
import psycopg2

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

QT_XLSX = r"C:\Users\R.Nozaki\Downloads\quotation_list.xlsx"
PJ_XLSX = r"C:\Users\R.Nozaki\Downloads\PJ_list_20251017_R0-1.xlsx"
DB = dict(host='localhost', port='5432', dbname='pegasus_erp', user='postgres', password='postgres')


def dval(v):
    """coerce numeric/date/str to safe value"""
    if v is None: return None
    if isinstance(v, str):
        s = v.strip()
        return s if s else None
    return v


def to_date(v):
    if v is None: return None
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    if isinstance(v, str):
        s = v.strip()
        if not s: return None
        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%b-%Y', '%b.%d,%Y', '%b %d,%Y'):
            try: return datetime.strptime(s, fmt).date()
            except ValueError: continue
    return None


def to_num(v):
    if v is None: return 0.0
    if isinstance(v, (int, float)): return float(v)
    s = str(v).replace(',', '').strip()
    try: return float(s)
    except: return 0.0


def norm_name(s):
    if not s: return ''
    s = str(s).lower()
    s = re.sub(r'[\(\),\.\-_/]', ' ', s)
    for w in ['co ltd','co.,ltd','company limited','limited','ltd','co','corporation',
              'corp','inc','pcl','public company','(thailand)','thailand','thai']:
        s = s.replace(w, ' ')
    return re.sub(r'\s+', ' ', s).strip()


def tokens(s):
    return set(t for t in norm_name(s).split() if len(t) >= 3)


def sim(a, b):
    ta, tb = tokens(a), tokens(b)
    if not ta or not tb: return 0.0
    return len(ta & tb) / max(len(ta), len(tb))


def resolve_customer(name, cust_list, cache):
    if not name: return None
    key = name.strip().lower()
    if key in cache: return cache[key]
    best_id, best = None, 0.0
    for cid, n in cust_list:
        s = sim(name, n)
        if s > best: best, best_id = s, cid
    result = best_id if best >= 0.3 else None
    cache[key] = result
    return result


def parse_due_date(remark):
    """Extract Due dd/mm/yyyy or similar from remark text."""
    if not remark: return None
    s = str(remark)
    m = re.search(r'Due\s*[:：]?\s*(\d{1,2})[/\-](\d{1,2})[/\-](\d{2,4})', s, re.I)
    if not m: return None
    d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
    if y < 100: y += 2000
    try: return date(y, mo, d)
    except: return None


def main():
    conn = psycopg2.connect(**DB)
    conn.autocommit = False
    cur = conn.cursor()

    # clear dependent tables first
    print('[0] Clearing existing project data...')
    cur.execute("UPDATE cost_sheets SET project_id = NULL WHERE project_id IS NOT NULL")
    cur.execute("UPDATE project_cost_items SET project_id = NULL WHERE project_id IS NOT NULL")
    cur.execute("UPDATE quotation_headers SET project_id = NULL WHERE project_id IS NOT NULL")
    cur.execute("DELETE FROM project_invoices")
    cur.execute("DELETE FROM project_purchases")
    cur.execute("DELETE FROM project_payment_schedules")
    cur.execute("DELETE FROM project_progress")
    cur.execute("DELETE FROM projects")
    conn.commit()

    # Load customers + POs
    cur.execute("SELECT customer_id, customer_name FROM customers WHERE is_deleted=FALSE AND customer_name IS NOT NULL")
    cust_list = cur.fetchall()
    cust_cache = {}

    cur.execute("SELECT po_id, po_no FROM purchase_order_headers WHERE is_deleted=FALSE")
    po_map = {r[1]: r[0] for r in cur.fetchall() if r[1]}
    print(f"  customers={len(cust_list)}  POs={len(po_map)}")

    # ---------- 1. Projects from Base ----------
    print('\n[1] Reading Base sheet...')
    wb_q = openpyxl.load_workbook(QT_XLSX, data_only=True, read_only=True)
    ws = wb_q['Base']
    projects = {}  # pj_no -> row dict
    for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if r_idx < 4: continue
        if len(row) < 35: row = tuple(list(row) + [None]*(35 - len(row)))
        pj_no = row[1]
        if not pj_no: continue
        pj_no = str(pj_no).strip()
        if not re.match(r'^PJ\d{6}$', pj_no): continue
        related = row[2]
        pj_segment = row[3]
        pj_category = row[4]
        pj_classification = row[5]
        pj_name = row[6] or pj_no
        customer = row[7]
        end_customer = row[8]
        total_revenue = to_num(row[9])
        cost_total = to_num(row[10])
        sales_hw = to_num(row[11])
        sales_sw = to_num(row[12])
        sales_swdev = to_num(row[13])
        sales_swlic = to_num(row[15])
        sales_inst = to_num(row[16])
        sales_swinst = to_num(row[17])
        sales_hwwir = to_num(row[18])
        total_cost_tc = to_num(row[19])
        profit = to_num(row[20])
        service_cost = to_num(row[21])
        engineer_cost = to_num(row[22])
        po_date = to_date(row[29])
        start_work = to_date(row[30])
        finished_work = to_date(row[31])
        plan_deliv = to_date(row[32])
        deliv = to_date(row[33])

        projects[pj_no] = dict(
            pj_no=pj_no,
            related_pj_no=str(related).strip() if related else None,
            pj_segment=str(pj_segment)[:200] if pj_segment else None,
            pj_category=str(pj_category)[:200] if pj_category else None,
            pj_classification=str(pj_classification)[:100] if pj_classification else None,
            pj_name=str(pj_name)[:500],
            customer=customer,
            end_user_customer=str(end_customer)[:300] if end_customer else None,
            total_revenue=total_revenue,
            cost_total=cost_total,
            sales_hardware=sales_hw, sales_software=sales_sw,
            sales_sw_development=sales_swdev, sales_sw_license=sales_swlic,
            sales_installation=sales_inst, sales_sw_installation=sales_swinst,
            sales_hw_wiring=sales_hwwir,
            total_cost=total_cost_tc, gross_profit=profit,
            service_cost=service_cost, engineer_cost=engineer_cost,
            po_date=po_date, start_work_date=start_work,
            finished_work_date=finished_work,
            plan_delivery_date=plan_deliv, delivery_date=deliv,
        )
    print(f'  Projects in Base: {len(projects)}')

    # ---------- 2. list sheet: ensure PJs referenced by quotations exist ----------
    print('\n[2] Reading list sheet for quotation PJ refs...')
    ws_l = wb_q['list']
    list_rows = []
    for r_idx, row in enumerate(ws_l.iter_rows(values_only=True), start=1):
        if r_idx < 3: continue
        if len(row) < 28: row = tuple(list(row) + [None]*(28 - len(row)))
        pj = row[3]
        if not pj: continue
        s = str(pj).strip()
        if not re.match(r'^PJ\d{6}$', s): continue
        list_rows.append((s, row))
    list_pjs = set(s for s, _ in list_rows)
    missing = list_pjs - set(projects.keys())
    print(f'  PJ Nos in list: {len(list_pjs)}  missing from Base: {len(missing)}')
    for s, row in list_rows:
        if s not in missing: continue
        projects[s] = dict(
            pj_no=s,
            related_pj_no=str(row[4]).strip() if row[4] else None,
            pj_segment=None,
            pj_category=str(row[10])[:200] if row[10] else None,
            pj_classification=None,
            pj_name=str(row[11])[:500] if row[11] else s,
            customer=row[6], end_user_customer=None,
            total_revenue=to_num(row[12]),
            cost_total=to_num(row[13]),
            sales_hardware=0, sales_software=0, sales_sw_development=0,
            sales_sw_license=0, sales_installation=0, sales_sw_installation=0,
            sales_hw_wiring=0, total_cost=to_num(row[13]),
            gross_profit=to_num(row[18]),
            service_cost=to_num(row[13]),
            engineer_cost=to_num(row[14]),
            po_date=to_date(row[5]),
            start_work_date=None, finished_work_date=None,
            plan_delivery_date=None, delivery_date=None,
        )
        missing.discard(s)
    print(f'  Total projects after merge: {len(projects)}')

    # ---------- 3. Insert projects ----------
    print('\n[3] Inserting projects...')
    inserted = 0
    pj_to_id = {}
    for pj_no, p in projects.items():
        cid = resolve_customer(p['customer'], cust_list, cust_cache)
        cur.execute("""
            INSERT INTO projects (
                pj_no, related_pj_no, pj_segment, pj_category, pj_classification,
                pj_name, customer_id, end_user_customer,
                total_revenue, cost_total,
                sales_hardware, sales_software, sales_sw_development,
                sales_sw_license, sales_installation, sales_sw_installation,
                sales_hw_wiring, total_cost, gross_profit,
                service_cost, engineer_cost,
                po_date, start_work_date, finished_work_date,
                plan_delivery_date, delivery_date, status
            ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'ACTIVE')
            RETURNING project_id
        """, [
            p['pj_no'], p['related_pj_no'], p['pj_segment'], p['pj_category'], p['pj_classification'],
            p['pj_name'], cid, p['end_user_customer'],
            p['total_revenue'], p['cost_total'],
            p['sales_hardware'], p['sales_software'], p['sales_sw_development'],
            p['sales_sw_license'], p['sales_installation'], p['sales_sw_installation'],
            p['sales_hw_wiring'], p['total_cost'], p['gross_profit'],
            p['service_cost'], p['engineer_cost'],
            p['po_date'], p['start_work_date'], p['finished_work_date'],
            p['plan_delivery_date'], p['delivery_date']
        ])
        pj_to_id[pj_no] = cur.fetchone()[0]
        inserted += 1
        if inserted % 100 == 0:
            conn.commit()
            print(f'  ...{inserted}')
    conn.commit()
    print(f'  Inserted {inserted} projects')

    # Link quotation_headers.project_id via pj_no
    cur.execute("""
        UPDATE quotation_headers qh
        SET project_id = p.project_id
        FROM projects p
        WHERE qh.pj_no = p.pj_no AND qh.project_id IS NULL
    """)
    linked = cur.rowcount
    conn.commit()
    print(f'  Linked {linked} quotations to projects via pj_no')

    # ---------- 4. PJ_list: invoices + purchases per sheet ----------
    print('\n[4] Reading PJ_list per-project sheets...')
    wb_pj = openpyxl.load_workbook(PJ_XLSX, data_only=True, read_only=True)
    total_inv = total_pur = sheets_done = no_match = 0
    for sheet_name in wb_pj.sheetnames:
        m = re.match(r'^(PJ\d{6})', sheet_name)
        if not m: continue
        pj_no = m.group(1)
        project_id = pj_to_id.get(pj_no)
        if not project_id:
            no_match += 1
            continue
        ws_p = wb_pj[sheet_name]
        # iterate rows once (read_only streams; ws.cell is O(n))
        inv_line = pur_line = 0
        for r_idx, row in enumerate(ws_p.iter_rows(values_only=True), start=1):
            if r_idx < 12: continue
            if r_idx > 250: break
            # pad row
            if len(row) < 18:
                row = tuple(list(row) + [None] * (18 - len(row)))
            # Invoices cols 1..7
            inv_date = row[1]; inv_no = row[2]; cust_n = row[3]
            amount = row[5]; remark = row[6]
            if inv_no or inv_date or amount:
                amt = to_num(amount)
                if amt or inv_no:
                    inv_line += 1
                    pay_d = parse_due_date(remark)
                    cur.execute("""
                        INSERT INTO project_invoices
                          (project_id, line_no, invoice_date, invoice_no, customer_name,
                           amount, remark, payment_date, payment_amount)
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    """, [project_id, inv_line, to_date(inv_date),
                          str(inv_no)[:100] if inv_no else None,
                          str(cust_n)[:300] if cust_n else None, amt,
                          str(remark)[:500] if remark else None,
                          pay_d, amt if pay_d else 0])
                    total_inv += 1
            # Purchases cols 8..18
            pdate = row[8]; pinv = row[9]; desc = row[10]
            pamt = row[12]; pay_terms = row[13]
            po_no = row[16]; pv_no = row[17]
            if po_no or pinv or pamt:
                amt = to_num(pamt)
                if amt or po_no or pinv:
                    pur_line += 1
                    po_id = po_map.get(str(po_no).strip()) if po_no else None
                    cur.execute("""
                        INSERT INTO project_purchases
                          (project_id, line_no, purchase_date, purchase_invoice_no,
                           description, amount, payment_terms, po_no, po_id, remark)
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    """, [project_id, pur_line, to_date(pdate),
                          str(pinv)[:100] if pinv else None,
                          str(desc)[:500] if desc else None,
                          amt,
                          str(pay_terms)[:200] if pay_terms else None,
                          str(po_no)[:100] if po_no else None, po_id,
                          str(pv_no)[:500] if pv_no else None])
                    total_pur += 1

        sheets_done += 1
        if sheets_done % 50 == 0:
            conn.commit()
            print(f'  ...{sheets_done} sheets  inv={total_inv} pur={total_pur}')

    conn.commit()
    # Roll up purchase totals into projects.purchase_actual
    cur.execute("""
        UPDATE projects p SET purchase_actual = COALESCE(s.total,0)
        FROM (SELECT project_id, SUM(amount) total FROM project_purchases
              WHERE is_deleted=FALSE GROUP BY project_id) s
        WHERE s.project_id = p.project_id
    """)
    conn.commit()

    print('\n=== Summary ===')
    print(f'  Projects inserted: {inserted}')
    print(f'  Quotations linked: {linked}')
    print(f'  PJ sheets processed: {sheets_done}')
    print(f'  PJ sheets no-match: {no_match}')
    print(f'  Invoices inserted: {total_inv}')
    print(f'  Purchases inserted: {total_pur}')
    cur.close(); conn.close()


if __name__ == '__main__':
    main()

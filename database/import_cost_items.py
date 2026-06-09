#!/usr/bin/env python3
"""
PEGASUS ERP - Import cost breakdown items from Excel to project_cost_items table.
Reads the '001_Breakdown cost' sheet and maps:
  D=category, E=description, F=supplier, G=brand, H=lead_time,
  J=unit_price, K=quantity, L=total, M=unit, N=remark
"""
import sys
import os
import openpyxl
import psycopg2

# Configuration
DB_DSN = "host=localhost dbname=pegasus_erp user=tomlogin password=Mpass4tomLogin"
SHEET_NAME = "001_Breakdown cost"
DATA_ROW_START = 9
DATA_ROW_END = 104  # inclusive (row 105 is Sub Total)

def import_cost_items(excel_path, project_id, user_id=1):
    """Import cost items from Excel breakdown sheet into project_cost_items."""

    if not os.path.exists(excel_path):
        print(f"ERROR: File not found: {excel_path}")
        return False

    wb = openpyxl.load_workbook(excel_path, data_only=True)

    if SHEET_NAME not in wb.sheetnames:
        # Try to find a sheet containing 'Breakdown' or 'cost'
        found = None
        for name in wb.sheetnames:
            if 'breakdown' in name.lower() or 'cost' in name.lower():
                found = name
                break
        if found:
            ws = wb[found]
            print(f"INFO: Using sheet '{found}' instead of '{SHEET_NAME}'")
        else:
            print(f"ERROR: Sheet '{SHEET_NAME}' not found. Available: {wb.sheetnames}")
            return False
    else:
        ws = wb[SHEET_NAME]

    conn = psycopg2.connect(DB_DSN)
    cur = conn.cursor()

    # Clear existing items for this project (from IMPORT source)
    cur.execute(
        "DELETE FROM project_cost_items WHERE project_id = %s AND source = 'IMPORT'",
        (project_id,)
    )
    print(f"Cleared existing IMPORT items for project_id={project_id}")

    current_category = None
    line_no = 0
    inserted = 0
    total_cost = 0

    for row in range(DATA_ROW_START, DATA_ROW_END + 1):
        d_val = ws.cell(row, 4).value   # D: Category/Model
        e_val = ws.cell(row, 5).value   # E: Description
        f_val = ws.cell(row, 6).value   # F: Supplier
        g_val = ws.cell(row, 7).value   # G: Brand
        h_val = ws.cell(row, 8).value   # H: Lead Time
        j_val = ws.cell(row, 10).value  # J: Unit price
        k_val = ws.cell(row, 11).value  # K: Quantity
        l_val = ws.cell(row, 12).value  # L: Total
        m_val = ws.cell(row, 13).value  # M: Unit
        n_val = ws.cell(row, 14).value  # N: Remark

        # Skip completely empty rows
        if not any([d_val, e_val, j_val, k_val, l_val]):
            continue

        # Clean up values
        d_str = str(d_val).strip() if d_val else None
        e_str = str(e_val).strip() if e_val else None
        f_str = str(f_val).strip() if f_val and str(f_val).strip() not in ['-', 'None'] else None
        g_str = str(g_val).strip() if g_val and str(g_val).strip() not in ['-', 'None'] else None
        h_str = str(h_val).strip() if h_val and str(h_val).strip() not in ['-', 'None'] else None
        m_str = str(m_val).strip() if m_val and str(m_val).strip() not in ['None'] else None
        n_str = str(n_val).strip() if n_val and str(n_val).strip() not in ['None'] else None

        # Determine if this is a category header row
        # Category rows: D column has a value AND (no unit price AND no quantity) OR it's a known category
        is_category = False
        if d_str:
            # Update current category
            current_category = d_str
            # Check if this is a pure category header (no price data) or a sub-header
            if (j_val is None or j_val == 0) and (k_val is None or k_val == 0) and (l_val is None or l_val == 0):
                is_category = True

        # Parse numeric values
        unit_price = 0
        quantity = 0
        total = 0
        try:
            if j_val is not None and isinstance(j_val, (int, float)):
                unit_price = round(float(j_val), 4)
        except (ValueError, TypeError):
            pass
        try:
            if k_val is not None and isinstance(k_val, (int, float)):
                quantity = round(float(k_val), 4)
        except (ValueError, TypeError):
            pass
        try:
            if l_val is not None and isinstance(l_val, (int, float)):
                total = round(float(l_val), 2)
        except (ValueError, TypeError):
            pass

        line_no += 1

        cur.execute("""
            INSERT INTO project_cost_items
                (project_id, line_no, category, description, supplier, brand, lead_time,
                 unit_price, quantity, total_amount, unit, remark, is_category_row, source, created_by)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 'IMPORT', %s)
        """, (
            project_id, line_no,
            current_category if not is_category else d_str,
            e_str if not is_category else d_str,
            f_str, g_str, h_str,
            unit_price, quantity, total, m_str, n_str,
            is_category, user_id
        ))
        inserted += 1
        total_cost += total

        status = "CAT " if is_category else "    "
        desc_show = (e_str or d_str or '(empty)')[:50]
        print(f"  {status} Row {row:3d} -> Line {line_no:3d}: {desc_show:50s} | {total:>14,.2f}")

    conn.commit()
    cur.close()
    conn.close()

    print(f"\n{'='*60}")
    print(f"Imported {inserted} items, Total cost: {total_cost:,.2f} Baht")
    print(f"Project ID: {project_id}")
    return True


if __name__ == '__main__':
    if len(sys.argv) < 3:
        print("Usage: python import_cost_items.py <excel_file> <project_id> [user_id]")
        sys.exit(1)

    excel_file = sys.argv[1]
    proj_id = int(sys.argv[2])
    uid = int(sys.argv[3]) if len(sys.argv) > 3 else 1

    success = import_cost_items(excel_file, proj_id, uid)
    sys.exit(0 if success else 1)

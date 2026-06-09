"""
Phase 2: Import all QT PDFs into quotation_headers + quotation_lines.

Inputs:
  - QT data/**/*.pdf       (2,624 TOMAS-issued quotation PDFs)
  - QT data/Quotation_List_Aggregated_by_AA.xlsx  (status + project link source)

Approach:
  pdftotext (poppler) -> regex parse:
    - Quotation No (header)
    - Date
    - Customer (To: ...)
    - Sub Total / VAT% / Discount / Grand Total
    - Line items (No 1-1 …)
  Insert quotation_headers (status mapped from Excel column U Possibility)
  Insert quotation_lines

Output:
  backups/qt_import_report.csv  (one row per PDF: pdf, qt_no, status, project_id, lines, total, error)
"""
from __future__ import annotations
import sys, os, re, json, csv, subprocess, datetime, glob
sys.stdout.reconfigure(encoding='utf-8')

import openpyxl, psycopg
from psycopg.rows import dict_row

ROOT = r"C:\Users\R.Nozaki\Downloads\Pegasus_ERP_R1"
PDF_ROOT = os.path.join(ROOT, "QT data")
XLSX     = os.path.join(PDF_ROOT, "Quotation_List_Aggregated_by_AA.xlsx")
REPORT   = os.path.join(ROOT, "backups", "qt_import_report.csv")

# Map Excel Possibility prefix -> quotation status
POSSIBILITY_TO_QT_STATUS = {
    "Deposited":  "WON",
    "Invoice":    "WON",
    "PO":         "WON",
    "A:Fix price & waiting PO": "APPROVED",
    "C:Considering,PO Next 1month": "SUBMITTED",
    "D:Quotation,PO Next 2months": "SUBMITTED",
    "E:Quotation,PO Next 3months or more": "DRAFT",
    "Revised":    "DRAFT",
    "Lost":       "LOST",
    "Pending":    "DRAFT",
}

POSSIBILITY_HEAD_RE = re.compile(r"^([^:]+)")

def norm_possibility(s) -> str:
    if not s: return ""
    m = POSSIBILITY_HEAD_RE.match(str(s).strip())
    return m.group(1).strip() if m else ""

def pdftotext(path: str) -> str:
    try:
        r = subprocess.run(
            ["pdftotext", "-layout", "-enc", "UTF-8", path, "-"],
            capture_output=True, timeout=30,
        )
        return r.stdout.decode("utf-8", errors="replace")
    except Exception as e:
        return ""

# Regexes for the standard TOMAS template
RE_QT_NO    = re.compile(r"Quotation\s+No\s*:\s*([A-Za-z0-9\-_/]+)", re.I)
RE_DATE     = re.compile(r"(?:Date|Issue\s*Date)\s*:\s*([0-9]{1,2}[-/][A-Za-z0-9]+[-/][0-9]{2,4})", re.I)
RE_TO       = re.compile(r"^\s*To\s*:\s+(.+?)$", re.M)
RE_SUBTOTAL = re.compile(r"Sub\s*Total\s+([0-9,]+\.\d{2})", re.I)
RE_DISCOUNT = re.compile(r"Discount\s+(-?[0-9,]+\.\d{2})", re.I)
RE_VAT_RATE = re.compile(r"Value\s*Added\s*Tax\s*([0-9.]+)\s*%", re.I)
RE_GRAND    = re.compile(r"Grand\s*Total(?:\s*Amount)?\s+([0-9,]+\.\d{2})", re.I)
RE_PROJECT  = re.compile(r"^\s*Project\s*:\s*(.+?)$", re.M)
RE_QUOTED_BY= re.compile(r"Quotation\s+by\s*:\s*(.+?)$", re.M)

# Line item: optional sub-numbering, then description then qty unit price ext_price
# Categories: "  1       Software" (no qty)
# Items:      "  1-1     Description                     1    job   480,000.00            480,000.00"
RE_CAT_ROW  = re.compile(r"^\s*(\d+)\s{3,}(.+?)\s*$")
RE_ITEM_ROW = re.compile(
    r"^\s*(\d+-\d+)\s+(.+?)\s+(\d+(?:\.\d+)?)\s+([A-Za-z]+)\s+([0-9,]+\.\d{2})\s+([0-9,]+\.\d{2})\s*$"
)
RE_ITEM_ROW_NOPRICE = re.compile(  # qty present but no unit/ext price (text-only sub-line)
    r"^\s*(\d+-\d+)\s+(.+?)\s+(\d+(?:\.\d+)?)\s+([A-Za-z]+)\s*$"
)

def parse_money(s: str) -> float:
    try:
        return float(s.replace(",", ""))
    except Exception:
        return 0.0

def parse_date(s: str) -> datetime.date | None:
    if not s: return None
    s = s.strip()
    for fmt in ("%d-%b-%y", "%d-%b-%Y", "%d/%m/%y", "%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.datetime.strptime(s, fmt).date()
        except Exception:
            pass
    return None

def parse_pdf(text: str) -> dict:
    out = {
        "quotation_no": None, "issue_date": None, "customer_text": None,
        "subtotal": 0.0, "discount": 0.0, "vat_rate": 7.0, "grand_total": 0.0,
        "project_text": None, "quoted_by": None,
        "lines": [],
    }
    if not text: return out

    m = RE_QT_NO.search(text);    out["quotation_no"] = m.group(1).strip() if m else None
    m = RE_DATE.search(text);     out["issue_date"]   = parse_date(m.group(1)) if m else None
    m = RE_TO.search(text);       out["customer_text"] = m.group(1).strip() if m else None
    m = RE_PROJECT.search(text);  out["project_text"]  = m.group(1).strip() if m else None
    m = RE_QUOTED_BY.search(text);out["quoted_by"]     = m.group(1).strip() if m else None
    m = RE_SUBTOTAL.search(text); out["subtotal"]      = parse_money(m.group(1)) if m else 0.0
    m = RE_DISCOUNT.search(text); out["discount"]      = parse_money(m.group(1)) if m else 0.0
    m = RE_VAT_RATE.search(text); out["vat_rate"]      = float(m.group(1)) if m else 7.0
    m = RE_GRAND.search(text);    out["grand_total"]   = parse_money(m.group(1)) if m else 0.0

    # Parse line items
    in_table = False
    parent_cat = None
    line_no = 0
    for raw_line in text.splitlines():
        if not in_table:
            if "Item / Description" in raw_line:
                in_table = True
            continue
        if any(k in raw_line for k in ("Sub Total", "Sub  Total", "Grand Total", "Sales Amount", "Value Added Tax")):
            break
        # Item row first
        m = RE_ITEM_ROW.match(raw_line)
        if m:
            line_no += 1
            out["lines"].append({
                "line_no": line_no,
                "parent_line_no": parent_cat,
                "is_category_row": False,
                "item_description": (m.group(1) + " " + m.group(2)).strip()[:500],
                "quantity": float(m.group(3)),
                "unit": m.group(4)[:20],
                "unit_price": parse_money(m.group(5)),
                "ext_price": parse_money(m.group(6)),
            })
            continue
        m = RE_ITEM_ROW_NOPRICE.match(raw_line)
        if m:
            line_no += 1
            out["lines"].append({
                "line_no": line_no,
                "parent_line_no": parent_cat,
                "is_category_row": False,
                "item_description": (m.group(1) + " " + m.group(2)).strip()[:500],
                "quantity": float(m.group(3)),
                "unit": m.group(4)[:20],
                "unit_price": 0.0,
                "ext_price": 0.0,
            })
            continue
        # Category row
        m = RE_CAT_ROW.match(raw_line)
        if m and m.group(2).strip() and not m.group(2).startswith("-"):
            line_no += 1
            parent_cat = line_no
            out["lines"].append({
                "line_no": line_no,
                "parent_line_no": None,
                "is_category_row": True,
                "item_description": m.group(2).strip()[:500],
                "quantity": 0.0, "unit": "", "unit_price": 0.0, "ext_price": 0.0,
            })
            continue
    return out


def load_excel_index() -> dict:
    """Return dict[ qt_no_normalised ] = {status, pj_no, customer, possibility} from Excel."""
    wb = openpyxl.load_workbook(XLSX, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    idx = {}
    for r in ws.iter_rows(min_row=3, values_only=True):
        if not r or len(r) < 21: continue
        qt = r[2]   # col C
        if not qt: continue
        qt_key = str(qt).strip().upper().replace("_", "-")
        idx[qt_key] = {
            "qt_no": str(qt).strip(),
            "pj_no": str(r[3]).strip() if r[3] else "",
            "customer": str(r[6]).strip() if r[6] else "",
            "possibility": str(r[20]).strip() if r[20] else "",
        }
    return idx


def qt_no_from_filename(fname: str) -> str | None:
    # QT250117_01_Pokayoke...pdf  →  QT250117-01
    base = os.path.basename(fname)
    m = re.match(r"(QT[\-_]?[0-9]+[\-_][0-9]+)", base, re.I)
    if not m: return None
    raw = m.group(1).upper().replace("_", "-")
    # Insert dash between digits if missing: QT250117-01
    return raw

def main():
    print("Loading Excel index...")
    excel_idx = load_excel_index()
    print(f"  Excel rows indexed: {len(excel_idx)}")

    print("Connecting to DB...")
    conn = psycopg.connect(
        host="localhost", port=5432, dbname="pegasus_erp",
        user="postgres", password=os.environ.get("PG_PASSWORD","postgres"),
        autocommit=False, row_factory=dict_row,
    )
    cur = conn.cursor()

    # ID maps
    div_id = cur.execute("SELECT division_id FROM divisions WHERE is_deleted=false ORDER BY division_id LIMIT 1").fetchone()["division_id"]
    admin_uid = cur.execute("SELECT user_id FROM users WHERE role='ADMIN' AND is_active=true ORDER BY user_id LIMIT 1").fetchone()["user_id"]
    proj_by_no = {r["pj_no"]: r["project_id"] for r in cur.execute("SELECT project_id, pj_no FROM projects")}
    def norm_name(n): return re.sub(r"[\s,\.\-_]+","", n or "").upper()
    cust_by_name = {norm_name(r["customer_name"]): r["customer_id"]
                    for r in cur.execute("SELECT customer_id, customer_name FROM customers WHERE is_deleted=false")}

    def get_or_create_customer(name: str) -> int:
        if not name: return None
        k = norm_name(name)
        if k in cust_by_name: return cust_by_name[k]
        code = f"CUS-IMP-{len(cust_by_name)+1:04d}"
        row = cur.execute(
            """INSERT INTO customers (customer_code, customer_name, division_id, country,
                  currency_code, payment_terms, approval_status,
                  created_by, submitted_by, submitted_at,
                  manager_approved_by, manager_approved_at, ceo_approved_by, ceo_approved_at)
               VALUES (%s,%s,%s,'TH','THB',30,'APPROVED',%s,%s,NOW(),%s,NOW(),%s,NOW())
               RETURNING customer_id""",
            (code, name[:200], div_id, admin_uid, admin_uid, admin_uid, admin_uid),
        ).fetchone()
        cust_by_name[k] = row["customer_id"]
        return row["customer_id"]

    pdfs = []
    for root, _, files in os.walk(PDF_ROOT):
        for f in files:
            if f.lower().endswith(".pdf") and not f.startswith("~$"):
                pdfs.append(os.path.join(root, f))
    print(f"  PDFs to process: {len(pdfs)}")

    report_rows = []
    stats = {"ok":0, "no_qt_no":0, "no_text":0, "duplicate":0, "error":0}
    seen_qt = set()

    for i, p in enumerate(pdfs, 1):
        try:
            text = pdftotext(p)
            if not text.strip():
                report_rows.append([p, "", "no_text", "", 0, 0, "pdftotext returned empty"])
                stats["no_text"] += 1
                continue

            parsed = parse_pdf(text)
            qt_no = parsed["quotation_no"]
            if not qt_no:
                # try filename
                qt_no = qt_no_from_filename(p)
            if not qt_no:
                report_rows.append([p, "", "no_qt_no", "", 0, 0, "no QT no in pdf or filename"])
                stats["no_qt_no"] += 1
                continue

            qt_norm = qt_no.upper().replace("_","-")
            if qt_norm in seen_qt:
                report_rows.append([p, qt_no, "duplicate", "", 0, 0, "duplicate QT no in this run"])
                stats["duplicate"] += 1
                continue
            seen_qt.add(qt_norm)

            # Look up Excel info
            info = excel_idx.get(qt_norm)
            poss_head = norm_possibility(info["possibility"]) if info else ""
            qt_status = POSSIBILITY_TO_QT_STATUS.get(poss_head, "DRAFT")
            project_id = proj_by_no.get(info["pj_no"]) if info and info.get("pj_no") else None

            # Customer: prefer Excel customer (more canonical) else PDF customer_text
            cust_name = (info["customer"] if info and info.get("customer") else parsed["customer_text"]) or ""
            cust_id = get_or_create_customer(cust_name) if cust_name else None
            if not cust_id:
                # cannot save without customer (NOT NULL) — skip
                report_rows.append([p, qt_no, "no_customer", info["pj_no"] if info else "", 0, parsed["grand_total"], "no customer name"])
                stats["error"] += 1
                continue

            # Calculate quotation totals
            subtotal_thb = parsed["subtotal"] or sum(l["ext_price"] for l in parsed["lines"] if not l["is_category_row"])
            vat_rate     = parsed["vat_rate"] or 7.0
            grand_total  = parsed["grand_total"] or (subtotal_thb * (1 + vat_rate/100))
            vat_amount   = max(0.0, grand_total - subtotal_thb)
            discount     = abs(parsed["discount"])
            issue_date   = parsed["issue_date"] or datetime.date.today()

            # Insert header
            qh = cur.execute(
                """
                INSERT INTO quotation_headers
                    (quotation_no, division_id, customer_id, project_name,
                     issue_date, currency_code, exchange_rate,
                     subtotal_thb, discount_amount, vat_rate, vat_amount, grand_total_thb,
                     status, quoted_by, created_by, created_at, updated_at)
                VALUES (%s,%s,%s,%s,%s,'THB',1,%s,%s,%s,%s,%s,%s,%s,%s,NOW(),NOW())
                RETURNING quotation_id
                """,
                (qt_no[:100], div_id, cust_id,
                 (parsed["project_text"] or "")[:300] or None,
                 issue_date, subtotal_thb, discount, vat_rate, vat_amount, grand_total,
                 qt_status, None, admin_uid),
            ).fetchone()
            qid = qh["quotation_id"]

            # Insert lines
            for ln in parsed["lines"]:
                cur.execute(
                    """
                    INSERT INTO quotation_lines
                      (quotation_id, line_no, parent_line_no, is_category_row,
                       item_description, quantity, unit, unit_price, discount_rate,
                       ext_price, cost_total, sort_order)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,0,%s,0,%s)
                    """,
                    (qid, ln["line_no"], ln["parent_line_no"], ln["is_category_row"],
                     ln["item_description"], ln["quantity"], ln["unit"],
                     ln["unit_price"], ln["ext_price"], ln["line_no"]),
                )

            # Link to project if matched
            if project_id and parsed["lines"]:
                # We don't have a direct quotation -> project FK here; the PJ link is
                # via the quotation_no being noted on the project. Leave for now.
                pass

            stats["ok"] += 1
            report_rows.append([p, qt_no, "ok", info["pj_no"] if info else "", len(parsed["lines"]), grand_total, ""])

            if stats["ok"] % 100 == 0:
                print(f"  ...{stats['ok']} QTs inserted (file {i}/{len(pdfs)})")
                conn.commit()

        except Exception as e:
            conn.rollback()
            report_rows.append([p, "", "error", "", 0, 0, str(e)[:200]])
            stats["error"] += 1

    conn.commit()
    print("\nDone.")
    for k,v in stats.items(): print(f"  {k:12s}: {v}")

    # Write report CSV
    os.makedirs(os.path.dirname(REPORT), exist_ok=True)
    with open(REPORT, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(["pdf_path","qt_no","status","pj_no","line_count","grand_total","error"])
        w.writerows(report_rows)
    print(f"Report: {REPORT}")
    conn.close()

if __name__ == "__main__":
    main()
